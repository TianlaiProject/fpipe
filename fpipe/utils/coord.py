import os
import astropy.units as u
from astropy.coordinates import AltAz, SkyCoord
from astropy.coordinates import EarthLocation
from astropy.coordinates.angles import Angle

from astropy.time import Time
import numpy as np

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

from openpyxl import load_workbook

_Lon = (106. + 51./60. + 24.0/3600.) * u.deg
_Lat = (25. + 39./60. + 10.6/3600.) * u.deg
_Location = EarthLocation.from_geodetic(_Lon, _Lat)
_dir_ = os.path.dirname(__file__)

def get_pointing_any_scan(time, alt0, az0, time_format='unix', feed_rotation=0,
        beam_pos_file = _dir_ + '/../data/beam_pos.dat'):
    
    '''
    estimate the pointing RA Dec accoriding obs time and pointing alt az 
    
    time: obs time
    '''

    #if alt0 > 90.: alt0=90.

    alt0 = np.array([alt0, ]).flatten()[:, None]
    az0  = np.array([az0,  ]).flatten()[:, None]

    alt0[alt0>90] = 90.


    za0  = 90. - alt0
    za0 = (za0 * u.deg).to(u.radian).value.astype('float64')
    az0 = (az0 * u.deg).to(u.radian).value.astype('float64')

    # position of 19 beam in unit or arcmin, from Wenkai's calculation
    # already rotated by 23.4 deg
    #x_position = np.array([  0.000,   5.263,   0.659,  -4.604,  -5.263,
    #                        -0.659,   4.604,  10.526,   5.922,   1.318,
    #                        -3.945,  -9.208,  -9.867, -10.526,  -5.922,
    #                        -1.318,   3.945,   9.208,   9.867], dtype='float64')
    #y_position = np.array([  0.000,  -2.277,  -5.6965, -3.419,   2.277,
    #                         5.6965,  3.419,  -4.555,  -7.974, -11.393,
    #                         -9.116, -6.838,  -1.142,   4.555,   7.974,
    #                         11.393,  9.116,   6.838,   1.142], dtype='float64')
    #y_position = - y_position
    #separation = np.sqrt(x_position ** 2 + y_position ** 2) * u.arcmin
    #position_angle  = np.arctan2(x_position, y_position) * u.rad + 23.4 * u.deg

    beam_pos = np.loadtxt(beam_pos_file)
    x_position = beam_pos[:,1]
    y_position = -beam_pos[:,2]
    separation = np.sqrt(x_position ** 2 + y_position ** 2) * u.arcmin
    position_angle  = np.arctan2(x_position, y_position) * u.rad 

    position_angle += feed_rotation * u.deg

    separation = separation[None, :]
    position_angle = position_angle[None, :]
    
    separation = separation.to(u.radian).value
    position_angle = position_angle.to(u.radian).value
    # make sure the feed always align E-W/N-S
    position_angle = position_angle + az0
    position_angle[position_angle > np.pi] -= 2. * np.pi
    position_angle[position_angle < -np.pi] += 2. * np.pi
    #for i in range(19):
    #    print '%2d'%(i+1), '%12.6f'%(position_angle[i] * 180./np.pi), \
    #          '%12.6f'%((position_angle[i] + az0) * 180./np.pi)
    
    sign = np.abs(position_angle)
    sign[sign==0] = np.inf
    sign = position_angle / sign
    
    position_angle[:, 0] = 0.
    separation[:, 0] = 0.

    za = np.cos(za0) * np.cos(separation) + np.sin(za0) * np.sin(separation) * np.cos(position_angle)
    za = np.arccos(za)
    
    #if za0 != 0:
    daz = (- np.cos(za0) * np.cos(za) + np.cos(separation)) / np.sin(za0)/ np.sin(za)
    
    daz[daz<-1] = -1
    daz[daz>1]  = 1
    #print daz.min(), daz.max()
    daz = np.arccos(daz) * sign
    #daz[za0[:, 0]==0, :] = (position_angle - 2. * az0)[za0[:, 0]==0, :]
    daz[za0[:, 0]==0, :] = (position_angle - az0)[za0[:, 0]==0, :]

    az = Angle(az0 + daz, u.radian) 
    za = Angle(za, u.radian)
    alt = 90.*u.deg - za
    
    #for i in range(19):
    #    print '%3d'%i, '%3dd%3dm%6.2fs'%alt[i].dms,  '%6dd%3dm%6.2fs'%az[i].dms
    
    
    _t, _alt, _az = np.broadcast_arrays(time[:, None], alt.deg, az.deg)
    _t = Time(_t, format=time_format, location=_Location)
    _alt = Angle(_alt, u.deg)
    _az  = Angle(_az,  u.deg)
    c0   = SkyCoord(alt=_alt, az=_az, frame='altaz', location=_Location, obstime=_t)
    c0   = c0.transform_to('icrs')
    

    return az.deg, alt.deg, c0.ra.deg, c0.dec.deg

def project_to_antenna_coord(alt, az, alt0=None, az0=None):
    
    if alt0 is None: alt0 = alt[0]
    if az0  is None: az0  = az[0]
    y = alt - alt0
    x = (az  - az0) * np.cos(alt * np.pi / 180.)
    
    return x, y

def _get_meridian_transit_time(obstime, location, target):
    
    altaz_frame = AltAz(obstime=obstime, location=location)
    
    target_altaz = target.transform_to(altaz_frame)
    close_points = np.abs(target_altaz.alt.deg - 90.).argmin()
    
    return obstime[close_points]

def beam_axes():
    fig2 = plt.figure(figsize=[12, 12])
    gs2  = gridspec.GridSpec(5 , 10, left=0.08, bottom=0.08, 
                             right=0.95, top=0.95, wspace=0.1, hspace=0.1)
    ax2 = []
    ax2.append(fig2.add_subplot(gs2[2, 4:6]))
    
    ax2.append(fig2.add_subplot(gs2[2, 2:4]))
    ax2.append(fig2.add_subplot(gs2[3, 3:5]))
    ax2.append(fig2.add_subplot(gs2[3, 5:7]))
    ax2.append(fig2.add_subplot(gs2[2, 6:8]))
    ax2.append(fig2.add_subplot(gs2[1, 5:7]))
    ax2.append(fig2.add_subplot(gs2[1, 3:5]))
    
    ax2.append(fig2.add_subplot(gs2[2, 0:2]))
    ax2.append(fig2.add_subplot(gs2[3, 1:3]))
    ax2.append(fig2.add_subplot(gs2[4, 2:4]))
    ax2.append(fig2.add_subplot(gs2[4, 4:6]))
    ax2.append(fig2.add_subplot(gs2[4, 6:8]))
    ax2.append(fig2.add_subplot(gs2[3, 7:9]))
    ax2.append(fig2.add_subplot(gs2[2, 8:10]))
    ax2.append(fig2.add_subplot(gs2[1, 7:9]))
    ax2.append(fig2.add_subplot(gs2[0, 6:8]))
    ax2.append(fig2.add_subplot(gs2[0, 4:6]))
    ax2.append(fig2.add_subplot(gs2[0, 2:4]))
    ax2.append(fig2.add_subplot(gs2[1, 1:3]))
    
    return fig2, ax2


def xyz2azalt(coord_file, min_row=2, max_row=None):
    '''

    convert the antenna xyz to az alt

    '''
    
    wb = load_workbook(coord_file)
    datasheet = wb[u'\u6d4b\u91cf\u6570\u636e']
    
    if max_row is None:
        max_row = datasheet.max_row
    time = [col for col in datasheet.iter_cols(min_col=1, max_col=1, 
                                               min_row=min_row, max_row=max_row,
                                               values_only=True)]
    time = np.array(time).flatten()
    time = Time(time) - 8. * u.hour # convert from Bejing time to UTC
    
    data_col_min = 20
    data_col_max = 22
    data_name = [col for col in datasheet.iter_cols(min_col=data_col_min, max_col=data_col_max, 
                                                    min_row=1, max_row=1,
                                                    values_only=True)]
    data = [col for col in datasheet.iter_cols(min_col=data_col_min, max_col=data_col_max, 
                                               min_row=min_row, max_row=max_row,
                                               values_only=True)]
    print 'read ant. coord %s %s %s'%(tuple([x[0] for x in data_name]))

    data = np.array(data)
    X = data[0]
    Y = data[1]
    Z = data[2]

    # alt = 90. - za
    sinalt = -Z / (X**2 + Y**2 + Z**2) ** 0.5
    alt = np.array(np.arcsin(sinalt) * 180. / np.pi)

    az = np.array(np.arctan2(Y, X) * 180. / np.pi)
    #az[az < 0] += 360
    az = 270. - az
    
    return time, az * u.deg, alt * u.deg

def egamma_to_cirs_ra(egamma_ra, time):
    '''
    http://reionization.org/wp-content/uploads/2013/03/HERA_Memo46_lst2ra.html
    '''
    from astropy import _erfa as erfa
    from astropy.coordinates.builtin_frames.utils import get_jd12

    era = erfa.era00(*get_jd12(time, 'ut1'))
    theta_earth = Angle(era, unit='rad')

    assert(isinstance(time, Time))
    gast = time.sidereal_time('apparent', longitude=0)
    cirs_ra = egamma_ra - (gast-theta_earth)
    return cirs_ra

def drift_azalt(time, drift_dec, time_format='unix', drift_mjd0=None, 
        force_meridian=True):

    '''

    Assuming meridian drift scan,
    get the az alt according to the init dec.

    time: obs time UTC, in the formate of time_format; 
    drift_dec: the dec setup at beginning.

    Note: time[0] MUST be start time targeting at drift_dec.
          Otherwise, dec is drifting due to the epoch differences.

    '''

    _t = Time(time, format=time_format, location=_Location)
    if drift_mjd0 is not None:
        t0 = Time(drift_mjd0, format='mjd', location=_Location)
    else:
        t0 = _t[0]
    ra0 = t0.sidereal_time('apparent')#.to(u.deg)
    ra0 = egamma_to_cirs_ra(ra0, t0)

    _s = SkyCoord(ra=ra0, dec=drift_dec)
    altaz_frame = AltAz(obstime=t0, location=_Location)

    _s = _s.transform_to(altaz_frame)
    az, alt = _s.az.deg, _s.alt.deg
    if force_meridian:
        'force pointing to the meridian'
        if (az < 90) * (az > -90) + (az > 270):
            az = 0.
        else:
            az = 180.

    _t, alt, az = np.broadcast_arrays(_t[:, None], alt, az)
    return _t, az * u.deg, alt * u.deg

